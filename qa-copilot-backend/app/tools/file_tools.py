"""账务文件生成与校验的纯业务引擎。"""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook

from app.core.constants import FileTemplateFormat
from app.exceptions import BadRequestException


def _convert_value(value: Any, field: dict[str, Any]) -> tuple[Any, str | None]:
    """按字段类型、精度、日期格式和码值映射转换单值。"""
    name = str(field["name"])
    required = bool(field.get("required", False))
    if value in (None, ""):
        if required:
            return value, f"{name} 为必填字段"
        return field.get("defaultValue"), None
    mapping = field.get("mapping", {})
    if str(value) in mapping:
        value = mapping[str(value)]
    data_type = str(field.get("dataType", "STRING"))
    try:
        if data_type == "INTEGER":
            converted: Any = int(value)
        elif data_type == "DECIMAL":
            converted = Decimal(str(value))
            precision = field.get("precision")
            if precision is not None:
                converted = converted.quantize(Decimal(1).scaleb(-int(precision)))
        elif data_type in {"DATE", "DATETIME"}:
            pattern = str(field.get("format") or ("%Y-%m-%d" if data_type == "DATE" else "%Y-%m-%d %H:%M:%S"))
            if isinstance(value, datetime | date):
                converted = value.strftime(pattern)
            else:
                converted = datetime.strptime(str(value), pattern).strftime(pattern)
        elif data_type == "BOOLEAN":
            normalized = str(value).strip().lower()
            if normalized not in {"true", "false", "1", "0", "yes", "no"}:
                raise ValueError
            converted = normalized in {"true", "1", "yes"}
        else:
            converted = str(value)
    except ValueError, TypeError, InvalidOperation:
        return value, f"{name} 无法转换为 {data_type}"
    length = field.get("length")
    if length is not None and len(str(converted)) > int(length):
        return converted, f"{name} 长度超过 {length}"
    return converted, None


def validate_records(
    records: list[dict[str, Any]], fields: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """逐行校验并返回规范化记录和可定位错误报告。"""
    normalized_records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_number, record in enumerate(records, start=1):
        normalized: dict[str, Any] = {}
        for field in fields:
            source_field = str(field["sourceField"])
            value, error = _convert_value(record.get(source_field), field)
            normalized[str(field["name"])] = value
            if error:
                errors.append(
                    {"row": row_number, "field": field["name"], "message": error, "value": record.get(source_field)}
                )
        normalized_records.append(normalized)
    return normalized_records, errors


def _format_fixed_width(value: Any, field: dict[str, Any]) -> str:
    text_value = "" if value is None else str(value)
    length = int(field.get("length") or len(text_value))
    char = str(field.get("paddingChar") or " ")[:1]
    return text_value.rjust(length, char) if field.get("padding") == "LEFT" else text_value.ljust(length, char)


def generate_file(template: object, records: list[dict[str, Any]]) -> tuple[bytes, str, str, dict[str, Any]]:
    """按模板生成 CSV、Excel、定长/分隔 TXT、JSON 或 XML 字节。"""
    normalized, errors = validate_records(records, template.fields)
    if errors:
        raise BadRequestException(f"文件数据校验失败，共 {len(errors)} 项错误")
    names = [str(field["name"]) for field in template.fields]
    file_format = FileTemplateFormat(template.file_format)
    encoding = "utf-8-sig" if template.encoding == "UTF-8" else "gbk"
    if file_format in {FileTemplateFormat.CSV, FileTemplateFormat.DELIMITED_TXT}:
        stream = io.StringIO(newline="")
        writer = csv.DictWriter(
            stream,
            fieldnames=names,
            delimiter=template.delimiter or ("," if file_format == FileTemplateFormat.CSV else "|"),
        )
        writer.writeheader()
        writer.writerows(normalized)
        content = stream.getvalue().encode(encoding)
        extension = "csv" if file_format == FileTemplateFormat.CSV else "txt"
        content_type = "text/csv" if extension == "csv" else "text/plain"
    elif file_format == FileTemplateFormat.FIXED_WIDTH_TXT:
        lines = [
            "".join(_format_fixed_width(record.get(str(field["name"])), field) for field in template.fields)
            for record in normalized
        ]
        content = ("\n".join(lines) + "\n").encode(encoding)
        extension, content_type = "txt", "text/plain"
    elif file_format == FileTemplateFormat.JSON:
        content = json.dumps(normalized, ensure_ascii=False, indent=2, default=str).encode(encoding)
        extension, content_type = "json", "application/json"
    elif file_format == FileTemplateFormat.XML:
        root = ElementTree.Element("records")
        for record in normalized:
            item = ElementTree.SubElement(root, "record")
            for name in names:
                ElementTree.SubElement(item, name).text = "" if record.get(name) is None else str(record[name])
        content = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        extension, content_type = "xml", "application/xml"
    else:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Data")
        sheet.append(names)
        for record in normalized:
            sheet.append([record.get(name) for name in names])
        output = io.BytesIO()
        workbook.save(output)
        content = output.getvalue()
        extension = "xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    total_field = template.trailer_config.get("totalAmountField")
    total_amount = (
        sum((Decimal(str(item.get(total_field) or 0)) for item in normalized), Decimal(0)) if total_field else None
    )
    return (
        content,
        extension,
        content_type,
        {
            "record_count": len(normalized),
            "total_amount": str(total_amount) if total_amount is not None else None,
            "validation_errors": [],
        },
    )


def parse_file(template: object, content: bytes) -> list[dict[str, Any]]:
    """按模板格式读取上传文件，供同一字段规则进行反向校验。"""
    file_format = FileTemplateFormat(template.file_format)
    encoding = "utf-8-sig" if template.encoding == "UTF-8" else "gbk"
    if file_format in {FileTemplateFormat.CSV, FileTemplateFormat.DELIMITED_TXT}:
        return list(
            csv.DictReader(
                io.StringIO(content.decode(encoding)),
                delimiter=template.delimiter or ("," if file_format == FileTemplateFormat.CSV else "|"),
            )
        )
    if file_format == FileTemplateFormat.FIXED_WIDTH_TXT:
        records: list[dict[str, Any]] = []
        for line in content.decode(encoding).splitlines():
            offset = 0
            record: dict[str, Any] = {}
            for field in template.fields:
                length = int(field.get("length") or 0)
                if length <= 0:
                    raise BadRequestException("定长模板的每个字段都必须配置 length")
                record[str(field["sourceField"])] = line[offset : offset + length].strip()
                offset += length
            records.append(record)
        return records
    if file_format == FileTemplateFormat.JSON:
        value = json.loads(content.decode(encoding))
        if not isinstance(value, list) or any(not isinstance(item, dict) for item in value):
            raise BadRequestException("JSON 文件必须是对象数组")
        return value
    if file_format == FileTemplateFormat.XML:
        root = ElementTree.fromstring(content)
        return [{child.tag: child.text for child in item} for item in root.findall("record")]
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]
